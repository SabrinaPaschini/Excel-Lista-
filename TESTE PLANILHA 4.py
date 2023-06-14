import openpyxl

#CRIAR UMA PLANILHA

book = openpyxl.Workbook()

#COMO VISUALIZAR PAGINAS EXISTENTES

print(book.sheetnames)

#COMO CRIAR UMA GUIA

book.create_sheet('nomes')

#COMO SELECIONAR UMA PÁGINA

#SELECIONANDO uma GUIA

nomes_page = book ['nomes']

nomes_page.append(['VAZIA','Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo'])

nomes_page.append(['Sabrina'])
nomes_page.append(['Ana'])
nomes_page.append(['Luana'])
nomes_page.append(['Fernanda'])

#SALVAR PLANILHA

book.save('planilha de nomes.xlsx')
